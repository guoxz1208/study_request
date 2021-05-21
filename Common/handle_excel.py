# coding = utf -8
"""
1-读取Excel文件
2-读取Excel表中sheet页
3-获取Excel有效行数、列数；
4-获取Excel某一行的值；
5-获取Excel某一列的值；
6-获取Excel某一个单元格内容
7-获取Excel全部内容
8-写入Excel文件内容
"""
import json

import xlrd
import xlwt
import openpyxl
import os
import sys
# base_path = os.getcwd()
base_path = os.path.abspath(os.path.join(os.getcwd(),'../'))
sys.path.append(base_path)

class HandExcel:
    # 读取Excel
    def read_excel(self,file_name=None):
        if file_name == None:
            excel_path = base_path + '/Data/登陆接口自动化测试用例.xlsx'
        else:
            excel_path = base_path + file_name
        excel_name = xlrd.open_workbook(excel_path)
        return excel_name

    # 读取Excel中sheet页
    def read_sheet(self,SheetName=None):
        if SheetName == None:
            SheetName = "Sheet1"
        excel_sheet = self.read_excel().sheet_by_name(SheetName)
        return excel_sheet

    # 获取Excel中有效总行数
    def get_row(self,SheetName=None):
        row_num = self.read_sheet(SheetName).nrows
        return row_num

    # 获取Excel中有效总列数
    def get_col(self,SheetName=None):
        col_num = self.read_sheet(SheetName).ncols
        return col_num

    # 获取Excel中某一行的值
    def get_row_value(self,SheetName=None,number=None):
        if number == None:
            number = 0
        row_value = self.read_sheet(SheetName).row_values(number)
        return row_value

    # 获取Excel中某一列的值
    def get_col_value(self,SheetName,number=None):
        if number==None:
            number = 0
        col_value = self.read_sheet(SheetName).col_values(number)
        return col_value

    # 获取Excel中某一个单元格的内容
    def get_cell_value(self,row_num,col_num,SheetName=None):
        cell_value = self.read_sheet(SheetName).cell_value(row_num,col_num)
        return cell_value

    # 写入excel文件中 date 数据，date是list数据类型， fields 表头
    def write_excel(self,row,col,data,index,file_name=None):
        if file_name == None:
            excel_path = base_path + '/Data/登陆接口自动化测试用例.xlsx'
        else:
            excel_path = base_path + file_name
        wb = openpyxl.load_workbook(excel_path)
        sheet_name = wb.sheetnames
        ws = wb[sheet_name[index]]
        # ws.title = sheetname
        ws.cell(row,col,data)
        wb.save(excel_path)


if __name__ == '__main__':
    hand_excel = HandExcel()
    # print(hand_excel.read_excel())
    print(hand_excel.read_sheet("test"))
    # print(hand_excel.get_row())
    # print(hand_excel.get_col())
    # print(hand_excel.get_row_value())
    # print(hand_excel.get_col_value(2))
    # print(hand_excel.get_cell_value(2,2))
    # data = {"uesrname":"admin"}
    # print(hand_excel.write_excel(10,1,json.dumps(data),1))